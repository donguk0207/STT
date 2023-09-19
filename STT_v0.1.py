import tkinter as tk
import speech_recognition as sr
import win32api
from tkinter import PhotoImage
from tkcalendar import DateEntry
from datetime import datetime
from openpyxl import Workbook, load_workbook


class VoiceToTextGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("찾아가는 민원 서비스")
        self.root.geometry("1280x1024")

        #음성녹음_영역
        record_title_label = tk.Label(self.root, text="◈◈◈음성 녹음 및 저장◈◈◈", font=("Helvetica", 24, "bold"), padx=10, pady=5)
        record_title_label.pack(pady=(20, 10))

        self.record_image = PhotoImage(file="mic1.png")  # 이미지 파일 경로를 입력하세요
        self.record_button = tk.Button(self.root, image=self.record_image, command=self.record_and_save)
        self.record_button.pack(pady=20, fill=tk.BOTH)

        self.previous_date = ""

        #프린터_영역
        printer_title_label = tk.Label(self.root, text="◈◈◈저장 결과 조회◈◈◈", font=("Helvetica", 24, "bold"))
        printer_title_label.pack(pady=(20, 10))

        self.calendar_label = tk.Label(self.root, text="날짜를 선택하세요:", font=("Helvetica", 14))
        self.calendar_label.pack(padx=20, pady=(10, 0))

        self.calendar = DateEntry(self.root, width=15, height=2, background='darkblue', foreground='white', borderwidth=2)
        self.calendar.pack(padx=20, pady=10)

        self.result_label = tk.Label(self.root, text="", wraplength=500, justify="center", font=("Helvetica", 12))
        self.result_label.pack(padx=20, pady=10)

        self.print_button = tk.Button(self.root, text="◈프린트 Click◈", command=self.print_excel, fg="red", relief="raised", width=15, height=3, font=("Helvetica", 18, "bold"))
        self.print_button.pack(pady=(0, 20))

        self.calendar.bind("<<DateEntrySelected>>", self.read_excel_and_print)

        #스크롤바영역
        self.result_frame = tk.Frame(self.root)
        self.result_frame.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)

        self.result_text = tk.Text(self.result_frame, wrap=tk.WORD, font=("Helvetica", 18))
        self.result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.result_scrollbar = tk.Scrollbar(self.result_frame, command=self.result_text.yview)
        self.result_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.result_text.config(yscrollcommand=self.result_scrollbar.set)

        self.refresh_calendar_and_excel_content()  # 캘린더 업데이트 및 엑셀 내용 출력 시작
        self.root.mainloop()

    def record_to_text(self):
        r = sr.Recognizer()
        with sr.Microphone() as source:
            print("###음성 녹음을 시작합니다...###")
            audio = r.listen(source)
            try:
                print("###음성을 텍스트로 변환 중...###")
                text = r.recognize_google(audio, language='ko-KR', show_all=False)
                return text
            except sr.UnknownValueError:
                print("음성을 인식할 수 없습니다.")
                return "음성 인식 오류"
            except sr.RequestError as e:
                print(f"음성을 변환하는 동안 오류가 발생했습니다: {e}")
                return "음성 변환 오류"

    def save_to_excel(self, text):
        current_date = datetime.now().strftime("%Y-%m-%d")
        if current_date != self.previous_date:
            self.previous_date = current_date
            print(f"새로운 날짜({current_date})입니다. 새로운 엑셀 파일을 생성 합니다.")

        filename = current_date + ".xlsx"
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["시간", "내용"])

        current_time = datetime.now()
        formatted_time = current_time.strftime("[%m월 %d일 %H시 %M분]")

        ws.append([formatted_time, text])
        wb.save(filename)

    def record_and_save(self):
        text = self.record_to_text()
        self.save_to_excel(text)
        print("저장 완료!")

    def read_excel_and_print(self, event):
        try:
            target_date = self.calendar.get_date()
            target_date_str = target_date.strftime("%Y-%m-%d")
            filename = target_date_str + ".xlsx"

            wb = load_workbook(filename)
            ws = wb.active

            content = "날짜: {}\n\n".format(target_date_str)
            content += "{:<20} {:<50}\n".format("시간", "내용")
            for row in ws.iter_rows(min_row=2, values_only=True):
                content += "{:<20} {:<50}\n".format(row[0], row[1])

            # 내용을 Text 위젯에 설정하고 스크롤을 맨 위로 이동
            self.result_text.delete(1.0, tk.END)  # 기존 내용 삭제
            self.result_text.insert(tk.END, content)
            self.result_text.yview_moveto(0)  # 스크롤을 맨 위로 이동
        except FileNotFoundError:
            self.result_text.delete(1.0, tk.END)  # 기존 내용 삭제
            self.result_text.insert(tk.END, "오늘날짜 녹음된 내용이 없습니다.")

    def print_excel(self):
        try:
            target_date = self.calendar.get_date()
            target_date_str = target_date.strftime("%Y-%m-%d")
            filename = target_date_str + ".xlsx"

            win32api.ShellExecute(0, "print", filename, None, ".", 0)
        except Exception as e:
            print("프린트 오류:", e)

    def refresh_calendar(self):
        self.calendar.set_date(datetime.now().date())  # 오늘 날짜로 캘린더 업데이트
        self.root.after(60000, self.refresh_calendar)  # 1분(60,000ms) 후에 다시 호출

    def refresh_calendar_and_excel_content(self):
        target_date = datetime.now().date()  # 오늘 날짜로 설정
        self.calendar.set_date(target_date)  # 캘린더 위젯 업데이트

        target_date_str = target_date.strftime("%Y-%m-%d")
        filename = target_date_str + ".xlsx"

        try:
            wb = load_workbook(filename)
            ws = wb.active

            content = "날짜: {}\n\n".format(target_date_str)
            content += "{:<20} {:<50}\n".format("시간", "내용")
            for row in ws.iter_rows(min_row=2, values_only=True):
                content += "{:<20} {:<50}\n".format(row[0], row[1])

            # 내용을 Text 위젯에 설정하고 스크롤을 맨 위로 이동
            self.result_text.delete(1.0, tk.END)  # 기존 내용 삭제
            self.result_text.insert(tk.END, content)
            self.result_text.yview_moveto(1)  # 스크롤을 맨 위로 이동
        except FileNotFoundError:
            self.result_text.delete(1.0, tk.END)  # 기존 내용 삭제
            self.result_text.insert(tk.END, "해당 날짜의 파일을 찾을 수 없습니다.")

        self.root.after(60000, self.refresh_calendar_and_excel_content)  # 1분(60,000ms) 후에 다시 호출

if __name__ == "__main__":
    root = tk.Tk()
    app = VoiceToTextGUI(root)
    root.mainloop()